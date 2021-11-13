import os
import json
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from xlsxreplacer import nameTrans, formatDialogue
from pretty import prettyJSON

if not os.path.exists("inputs"):
	os.mkdir("inputs")
if not os.path.exists("outputs"):
	os.mkdir("outputs")

files = os.listdir("inputs")
verNum = "1.0.0" # Update this with new releases!!!


print("Parsing files...")

for item in files:
	#grab original text
	ref = open("inputs/" + item, encoding="utf-8")
	parsed = ref.read()
	ref.close()
	
	#prettified output (for debugging/reference)
	#prettyJSON(parsed, item)
	
	try:
		dict = json.loads(parsed)#converts json to dict format
	except json.decoder.JSONDecodeError:
		print("> Didn't parse " + item + " (file not in JSON format)")
	except:
		print("> Didn't parse " + item + " (failed to read JSON file)")
	else:
		newname = "outputs/" + item[0:item.find(".")] + "_parsed.xlsx"
		if os.path.exists(newname):#deletes _parsed file if it already exists
			os.remove(newname)
		new_file = openpyxl.Workbook()
		sheet=new_file.active
		sheet.title = item[0:item.find(".")]
		
		for group in dict:
			if group.startswith("voice"):
				sheet.append([group, dict[group]["serif"].replace("@", "\n")])#voice line jasons are so simple. bless
			
			
			elif group.startswith("group"):
				sheet.append(["GROUP", group])
				for x in dict[group]:
					for y, z in x.items(): #i don't know if there's a way to do this without three goddamn levels of nesting no. twst's file formatting sucks
						if y == "serif" and "text" in z.keys() and "speaker" in z.keys():# handles main dialogue (serifs)
							if "visible" in z["speaker"].keys() and not z["speaker"]["visible"]: # handles serifs w/o a speaker
								sheet.append(["SFX", formatDialogue(z["text"])])
							elif "callNext" in z.keys() and not z["callNext"]: # handles bits of dialogue separated by other actions of some sort
								sheet.append([nameTrans(z["speaker"]["text"]), formatDialogue(z["text"])])
							elif "callClear" in z.keys() and not z["callClear"]: # other half of the callNext code
								sheet["B"+str(sheet._current_row)].value=sheet["B"+str(sheet._current_row)].value + formatDialogue(z["text"])
							else:
								sheet.append([nameTrans(z["speaker"]["text"]), formatDialogue(z["text"])])
						elif y == "serif" and not "text" in z.keys(): # skips serifs without text
							pass
						
						elif y == "choice": # handles choices
							count = 1
							for obj in z:
								sheet.append(["[CHOICE]", str(count) + ": " + formatDialogue(obj["text"]), "GOTO: " + obj["goTo"]])
								count+=1
						
						elif y == "goTo":
							sheet.append(["[GOTO]", z["goTo"]]) # handles pathing
						
						elif y == "title" and "textWhite" in z.keys() and "textGold" in z.keys(): # handles titles
							if "episode" in z.keys(): # handles titles for episodes
								sheet.append(["[TITLE]", "Episode " + z["episode"]["storyNum"] + ": " + z["textWhite"] + "\n" + z["textGold"]])
							elif "prologue" in z.keys(): # handles titles for the prologue
								sheet.append(["[TITLE]", "Prologue " + z["prologue"]["storyNum"] + ": " + z["textWhite"] + "\n" + z["textGold"]])
							elif "eventEpisode" in z.keys(): # handles titles for events
								sheet.append(["[TITLE]", "Event Episode " + z["eventEpisode"]["storyNum"] + ": " + z["textWhite"] + "\n" + z["textGold"]])
							elif "eventTitle" in z.keys(): # ...also handles titles for events
								sheet.append(["[TITLE]", "Event Episode " + z["eventTitle"]["storyNum"] + ": " + z["textWhite"] + "\n" + z["textGold"]])
							elif "personal" in z.keys(): # handles titles for personal stories
								sheet.append(["[TITLE]", z["textGold"] + " " + z["personal"]["rarity"] + ": " + z["personal"]["anotherName"] + "\n" + z["textWhite"]])
							elif "talk" in z.keys(): # handles titles for chats
								sheet.append(["[TITLE]", z["textGold"] + " Chat: " + z["talk"]["text"] + "\n" + z["textWhite"]])
							else: # handles other titles
								sheet.append(["[TITLE]", z["textWhite"] + "\n" + z["textGold"]])
						elif y == "title" and not "textWhite" in z.keys(): # skips titles without text
							pass
						
						elif y == "text" and "text" in z.keys(): # handles generic "text" content
							sheet.append(["[TEXT]", formatDialogue(z["text"])])
						
						elif y == "place":
							if "jp" in z.keys() and "en" in z.keys(): # handles place titles
								sheet.append(["[PLACE]", z["jp"] + "\n" + z["en"]])
							elif not "jp" in z.keys(): # skips places without text
								pass
							
						elif y == "narration":
							if "text" in z.keys(): # handles narration boxes
								sheet.append(["[NARRATION]", z["text"]])
							elif "fadeOut" in z.keys():
								pass
							elif "delete" in z.keys() and z["delete"]:
								pass
						
						elif y == "telop":
							if "text" in z.keys(): #handles "telops"
								sheet.append(["[TELOP]", z["text"]])
							elif not "text" in z.keys(): # skips telops without text
								pass
						
						elif y == "balloon":
							if "text" in z.keys() and "speaker" in z.keys(): # handles speech balloons (in rhythmics and battle cutscenes)
								sheet.append([nameTrans(z["speaker"]), formatDialogue(z["text"])])
							elif not "text" in z.keys(): # skips balloons without text
								pass
						
						elif y == "cardSprite":
							if "cardId" in z.keys(): # handles card graphics
								if "delete" in z.keys() and z["delete"]:
									sheet.append(["[CARD]", "Hide Card " + str(z["cardId"])])
								elif "isGroovy" in z.keys() and z["isGroovy"]:
									sheet.append(["[CARD]", "Show Card " + str(z["cardId"]) + " Groovy"])
								else:
									sheet.append(["[CARD]", "Show Card " + str(z["cardId"])])
							elif "delete" in z.keys() and z["delete"] == True:
								sheet.append(["[CARD]", "Hide Card"])
						
						elif y == "sprite":
							if "path" in z.keys(): # handles item sprites
								sheet.append(["[SPRITE]", "Show " + z["id"] + " (" + z["path"][z["path"].rfind("/") + 1:] + ")"])
							elif "fadeOut" in z.keys():
								sheet.append(["[SPRITE]", "Fade out " + z["id"]])
							elif "visible" in z.keys() and z["visible"] == True:
								sheet.append(["[SPRITE]", "Show " + z["id"]])
							elif "visible" in z.keys() and z["visible"] == False:
								sheet.append(["[SPRITE]", "Hide " + z["id"]])
							elif "delete" in z.keys() and z["delete"]:
								sheet.append(["[SPRITE]", "Hide " + z["id"]])
						
						elif y == "movie" and "path" in z.keys(): # handles movies
							sheet.append(["[MOVIE]", z["path"]])
						elif y == "mirrorMovie":
							if "delete" in z.keys() and z["delete"] == True: # handles movies but for the mirrors
								pass
							elif "mirrorId" in z.keys() and not "animation" in z.keys():
								sheet.append(["[MIRROR]", "Movie " + str(z["mirrorId"])])
							else:
								pass
						
						elif y == "blot":
							if "dormitoryId" in z.keys() and "animation" in z.keys() and "phase" in z["animation"].keys(): # handles (over)blot animations
								sheet.append(["[BLOT]", "Dorm " + str(z["dormitoryId"]) + " Phase " + str(z["animation"]["phase"])])
							elif "animation" in z.keys() and "phase" in z["animation"].keys():
								sheet.append(["[BLOT]", "Phase " + str(z["animation"]["phase"])])
							elif "dormitoryId" in z.keys() and not "animation" in z.keys():
								pass
							elif "delete" in z.keys() and z["delete"]:
								pass
						elif y == "overBlot":
							if "dormitoryId" in z.keys() and "animation" in z.keys():
								sheet.append(["[BLOT]", "Dorm " + str(z["dormitoryId"]) + " Overblot"])
							elif "dormitoryId" in z.keys() and not "animation" in z.keys():
								pass
							elif "delete" in z.keys() and z["delete"]:
								pass
						
						elif y == "live2d" or y == "moveCamera" or y == "systemUI" or y == "run" or y == "bgm" or y == "advBgOperator" or y == "touch" or y == "wait" or y == "zoomCamera" or y == "se" or y == "curtain" or y == "spine" or y == "spineCharacter" or y == "sd" or y == "spfx" or y == "shakeCamera" or y == "voice" or y == "transition" or y == "spriteUI" or y == "emotion" or y == "vibration" or y == "advBg" or y == "shakeLoopCamera": # skips a bunch of boring animation/visual logicistical code
							pass
						
						else: # for debugging mostly and to catch types i missed
							sheet.append([y, "(no code to handle this type of object yet, sorry! --Ylime)"])
			
			
			elif group.startswith("word") or group.startswith("w") or group.startswith("cut") or group.startswith("c") or group.startswith("motion"): # this applies to both the intro text and also rhythmics. yes <3
				groupPrinted = False; 
				for x in dict[group]:
					if(groupPrinted == False and ("balloon" in x.keys() or "narration" in x.keys())):
						sheet.append(["[" + group + "]"]);
						groupPrinted = True;
					for y, z in x.items():
						if y == "balloon":
							if "text" in z.keys() and "targetId" in z.keys():
								sheet.append([z["targetId"], formatDialogue(z["text"])])
							elif "delete" in z.keys() and z["delete"]:
								pass
						
						elif y == "narration":
							if "text" in z.keys():
								sheet.append(["[NARRATION]", z["text"]])
							elif not "text" in z.keys():
								pass
						
						elif y == "voice" or y == "voiceWait" or y == "wait" or y == "spineCharacter" or y == "moveCamera" or y == "zoomCamera" or y == "emotion" or y == "spfxTriggerKicker" or y == "spfxTargetPointFollower" or y == "spine":
							pass
						
						else:
							sheet.append([y, str(z) + " (no code to handle this type of object yet, sorry! --Ylime)"])
			
			elif group.startswith("initialize"): # skips past rhythmic animations
				sheet.append(["(no dialogue in this type of JSON object!)"])
				break
			
			else:
				sheet.append(["(no code to handle this type of JSON file yet, sorry! --Ylime)"])

		sheet.append(["TwstStoryReader v" + verNum + " (Excel Build) by Ylimegirl\nhttps://github.com/Ylimegirl/TwstStoryReader"])
		
		sheet.column_dimensions["A"].width=15
		sheet.column_dimensions["B"].width=75
		for cell in sheet["A"]:
			cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
		for cell in sheet["B"]:
			cell.alignment = Alignment(wrap_text=True, vertical="center")
		if(sheet.max_column >= 3):
			sheet.column_dimensions["C"].width=15
			for cell in sheet["C"]:
				cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
		sheet.merge_cells(start_row=sheet.max_row, start_column=1, end_row=sheet.max_row, end_column=sheet.max_column)
		sheet["A"+str(sheet.max_row)].alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
		sheet.row_dimensions[sheet.max_row].height=30
		new_file.save(filename=newname)
		
		print("> Parsed " + item)

print("Finished parsing all files.")